import streamlit as st
import pandas as pd
import re
import csv
from io import BytesIO
import tempfile
import gc
from openpyxl import Workbook

# =========================
# CONFIG
# =========================
MAX_ROWS_PER_SHEET = 999000

st.set_page_config(page_title="BOP Usage List", layout="wide")

st.title("📊 BOP HU-WHERE-USED")
st.caption("⚡ Handles 1M+ rows | Multi-file supported | Streaming mode")

# =========================
# DATA STRUCTURE
# =========================
ALL_COLUMNS = [
    "Level","Search Object (SO)","Description DC","Item Quantity DU","Direct Usage","Description DU",
    "Status DU (MMA/DOC)","Plant status DU (MMA)","Status DU (BOM)","System Desc. DU","Plant DU (BOM)",
    "Plant Name DU","Final Usage","Description FU","Status FU(MMA/DOC)","Plant status FU (MMA)",
    "Status FU (BOM)","System Description FU","Plant FU (BOM)","FU Charact. 1 Value","Subitem Number",
    "Install. Point","Sub Item Quantity","Valid from CMA DU","Valid from date DU","Valid to CMA DU",
    "Valid to date DU","Status Text","WU Status Code"
]

GROUP_COLUMNS = [
    "Group","Part Number","Level","Description DC","Item Quantity DU","Direct Usage",
    "Final Usage","Description FU","Plant FU (BOM)","FU Charact. 1 Value"
]

# =========================
# FUNCTIONS
# =========================
def normalize_number(v):
    return re.sub(r"\D", "", v or "")

def identify_group(fu):
    fu = normalize_number(fu)
    if fu.startswith(("7612","7609","764","750","751","752")):
        return "CP1"
    if fu.startswith("0263"):
        return "CP2"
    if fu.startswith(("7620","7607")):
        return "CP1-PRO"
    if fu.startswith("8613600"):
        return "Bombardier"
    if fu.startswith("1270020"):
        return "E-bike"
    return "Other"

def parse_line(line, ncols):
    cols = line.strip().split("\t")
    if len(cols) == 1:
        cols = re.split(r'(?<!\S)\s{2,}(?!\S)', line.strip())
    if len(cols) < ncols:
        cols += [""] * (ncols - len(cols))
    return cols[:ncols]

def clean_cell(value):
    if value is None:
        return ""
    value = re.sub(r"[\x00-\x1F\x7F]", "", str(value))
    value = value.replace("\n", " ").replace("\r", " ")
    return value[:32767]

# =========================
# UI
# =========================
st.subheader("📂 Upload TXT files")

files = st.file_uploader(
    "Drag & drop TXT files here",
    type=["txt"],
    accept_multiple_files=True
)

run = st.button("🚀 Process")
reset = st.button("🔄 Reset")

# =========================
# SESSION STATE
# =========================
if "df_group" not in st.session_state:
    st.session_state.df_group = None

if "csv_path" not in st.session_state:
    st.session_state.csv_path = None

if "total_rows" not in st.session_state:
    st.session_state.total_rows = 0

if "excel_ready" not in st.session_state:
    st.session_state.excel_ready = False

if "excel_data" not in st.session_state:
    st.session_state.excel_data = None

if "group_filter" not in st.session_state:
    st.session_state.group_filter = []

if "search" not in st.session_state:
    st.session_state.search = ""

# =========================
# RESET
# =========================
if reset:
    st.session_state.df_group = None
    st.session_state.csv_path = None
    st.session_state.total_rows = 0
    st.session_state.excel_ready = False
    st.session_state.excel_data = None
    st.session_state.group_filter = []
    st.session_state.search = ""
    st.rerun()

if not files:
    st.info("📂 Upload TXT file(s)")
    st.stop()

# =========================
# PROCESS FILES
# =========================
if run:

    st.session_state.excel_ready = False
    st.session_state.excel_data = None

    progress = st.progress(0)
    status = st.empty()

    status.info("📥 Processing files...")

    tmp = tempfile.NamedTemporaryFile(delete=False, mode="w", newline="", encoding="utf-8")
    writer = csv.writer(tmp)
    writer.writerow(ALL_COLUMNS)

    group_rows = []
    total = sum(len(f.getvalue().decode("utf-8", errors="ignore").splitlines()) for f in files)
    processed = 0

    for f in files:
        text = f.getvalue().decode("utf-8", errors="ignore")

        for line in text.splitlines():

            if not line or line.startswith("Level"):
                continue

            cols = parse_line(line, len(ALL_COLUMNS))
            cols = (cols + [""] * len(ALL_COLUMNS))[:len(ALL_COLUMNS)]

            writer.writerow(cols)

            group_rows.append([
                identify_group(cols[12]),
                cols[12],
                cols[0],
                cols[2],
                cols[3],
                cols[4],
                cols[12],
                cols[13],
                cols[18],
                cols[19],
            ])

            processed += 1

            if processed % 5000 == 0:
                progress.progress(min(processed / total, 1.0))
                status.text(f"Processing {processed:,} rows")

    tmp.close()

    st.session_state.df_group = pd.DataFrame(group_rows, columns=GROUP_COLUMNS)
    st.session_state.csv_path = tmp.name
    st.session_state.total_rows = processed

    status.success("✅ Processing completed")

# =========================
# LOAD DATA
# =========================
if st.session_state.df_group is None:
    st.warning("⚠️ Click PROCESS")
    st.stop()

df_group = st.session_state.df_group

# =========================
# GENERATE EXCEL ONCE
# =========================
st.divider()

if not st.session_state.excel_ready:

    status = st.empty()
    progress_excel = st.progress(0)

    status.info("📦 Preparing Excel...")

    excel_buffer = BytesIO()
    wb = Workbook(write_only=True)

    ws = wb.create_sheet("Usage_1")
    ws.append(ALL_COLUMNS)

    total_rows = st.session_state.total_rows
    count = 0
    sheet_count = 1

    with open(st.session_state.csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)

        for row in reader:

            row = (row + [""] * len(ALL_COLUMNS))[:len(ALL_COLUMNS)]
            row = [clean_cell(x) for x in row]

            if count > 0 and count % MAX_ROWS_PER_SHEET == 0:
                sheet_count += 1
                ws = wb.create_sheet(f"Usage_{sheet_count}")
                ws.append(ALL_COLUMNS)

            ws.append(row)
            count += 1

            if count % 20000 == 0:
                progress_excel.progress(min(count / total_rows, 1.0))
                status.text(f"Writing {count:,} rows")

    progress_excel.progress(1.0)

    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.session_state.excel_data = excel_buffer.getvalue()
    st.session_state.excel_ready = True

    status.success(f"✅ Excel ready ({sheet_count} sheets)")

# =========================
# DOWNLOAD EXCEL
# =========================
if st.session_state.excel_ready:

    st.download_button(
        "⬇️ Download FULL Excel (Usage List)",
        st.session_state.excel_data,
        "BOP_Usage_List.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# FILTERS + DASHBOARD
# =========================
st.divider()

group_filter = st.multiselect(
    "Filter Group",
    sorted(df_group["Group"].unique()),
    default=st.session_state.group_filter
)

search = st.text_input("Search", value=st.session_state.search)

st.session_state.group_filter = group_filter
st.session_state.search = search

df_view = df_group.copy()

if group_filter:
    df_view = df_view[df_view["Group"].isin(group_filter)]

if search:
    df_view = df_view[df_view.astype(str).apply(
        lambda x: x.str.contains(search, case=False, na=False)
    ).any(axis=1)]

c1, c2, c3 = st.columns(3)
c1.metric("Total Rows", len(df_group))
c2.metric("Filtered", len(df_view))
c3.metric("Groups", df_group["Group"].nunique())

st.dataframe(df_view, width="stretch", height=500)

gc.collect()
